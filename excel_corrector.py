import pandas as pd
from openpyxl import load_workbook

from config import BASE_DATA_LIST_NAME, YANDEX_DISC_PIVOT_TABLE_FILE_NAME, MAIN_DIR, REMOTE_PATH, LOCAL_PIVOT_TABLE_PATH
from yandex_disk import download_file, upload_file


class ExcelCorrector:
    @staticmethod
    def correct_excel(data):
        download_file(REMOTE_PATH + YANDEX_DISC_PIVOT_TABLE_FILE_NAME, LOCAL_PIVOT_TABLE_PATH)
        wb = load_workbook(f'{MAIN_DIR}{YANDEX_DISC_PIVOT_TABLE_FILE_NAME}')
        ws = wb[BASE_DATA_LIST_NAME]

        # удаляем старые данные
        ws.delete_rows(2, ws.max_row)

        # записываем новые
        for row in data.itertuples(index=False):
            ws.append(list(row))

        wb.save(f'{MAIN_DIR}{YANDEX_DISC_PIVOT_TABLE_FILE_NAME}')
        upload_file(f'{MAIN_DIR}{YANDEX_DISC_PIVOT_TABLE_FILE_NAME}', REMOTE_PATH + YANDEX_DISC_PIVOT_TABLE_FILE_NAME)